from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from utils.db_utils import query_db, get_db
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from flask import send_file
import io
programme_config_bp = Blueprint("programme_config", __name__, url_prefix="/programme_config")

# ============================================================== #
# 🔹 LISTE DES PROGRAMMES
# ============================================================== #
@programme_config_bp.route("/")
def liste_programmes():
    # 🔍 Recherche
    q = (request.args.get("q", "") or "").strip()

    # 📄 Pagination
    page = request.args.get("page", 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page

    # 🔹 Base query
    sql_base = "FROM Programme WHERE 1=1"
    params = []

    if q:
        sql_base += " AND nom LIKE ? COLLATE NOCASE"
        params.append(f"%{q}%")

    # 🔢 Total pour pagination
    total_row = query_db(f"SELECT COUNT(*) as count {sql_base}", params, one=True)
    total = total_row["count"] if total_row else 0
    total_pages = (total + per_page - 1) // per_page  # arrondi supérieur

    # 📋 Données paginées
    sql = f"""
        SELECT id, nom, type
        {sql_base}
        ORDER BY id DESC
        LIMIT ? OFFSET ?
    """
    params += [per_page, offset]
    programmes = [dict(r) for r in query_db(sql, params)]

    # 🧭 Rendu template
    return render_template(
        "Programme/programme_config_liste.html",
        programmes=programmes,
        q=q,
        page=page,
        total_pages=total_pages,
        total=total
    )


# ============================================================== #
# 🔹 AJOUT / SUPPRESSION PROGRAMME
# ============================================================== #
@programme_config_bp.route("/ajouter", methods=["POST"])
def ajouter_programme():
    nom = request.form.get("nom")
    if not nom:
        flash("⚠️ Le nom du programme est obligatoire.", "warning")
        return redirect(url_for("programme_config.liste_programmes"))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO Programme (nom,type, idate, iuser)
        VALUES (?,?, DATETIME('now'), ?)
    """, (nom, session.get('user', {}).get('username', 'admin')))
    conn.commit()
    flash("✅ Programme ajouté avec succès.", "success")
    return redirect(url_for("programme_config.liste_programmes"))


@programme_config_bp.route("/supprimer/<int:id>", methods=["POST"])
def supprimer_programme(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM Programme WHERE id = ?", (id,))
    conn.commit()
    flash("🗑️ Programme supprimé avec succès.", "success")
    return redirect(url_for("programme_config.liste_programmes"))


# ============================================================== #
# 🔹 PHASES
# ============================================================== #
@programme_config_bp.route("/get_phases/<int:programme_id>")
def get_phases(programme_id):
    phases = query_db("""
        SELECT ph.id, f.nom AS nom_phase, ph.poids
        FROM programme_phase ph
        JOIN Phase f ON f.id = ph.phase_id
        WHERE ph.programme_id = ?
        ORDER BY f.id
    """, [programme_id])
    return jsonify([dict(r) for r in phases])


@programme_config_bp.route("/get_available_phases/<int:programme_id>")
def get_available_phases(programme_id):
    phases = query_db("""
        SELECT f.id, f.nom
        FROM Phase f
        WHERE f.id NOT IN (
            SELECT phase_id FROM programme_phase WHERE programme_id = ?
        )
        ORDER BY f.id
    """, [programme_id])
    return jsonify([dict(r) for r in phases])


@programme_config_bp.route("/ajouter_phase", methods=["POST"])
def ajouter_phase():
    programme_id = request.form.get("programme_id")
    phase_id = request.form.get("phase_id")
    poids = float(request.form.get("poids", 0))

    total = query_db("SELECT SUM(poids) AS total FROM programme_phase WHERE programme_id = ?", [programme_id], one=True)
    total_poids = (total["total"] or 0) + poids
    if total_poids > 100:
        return jsonify({"error": f"La somme totale ({total_poids:.1f}%) dépasse 100%."}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO programme_phase (programme_id, phase_id, poids, iuser)
        VALUES (?, ?, ?, ?)
    """, (programme_id, phase_id, poids, session.get('user', {}).get('username', 'admin')))
    conn.commit()
    return jsonify({"success": True})


@programme_config_bp.route("/modifier_phase/<int:id>", methods=["POST"])
def modifier_phase(id):
    poids = float(request.form.get("poids", 0))
    phase_info = query_db("SELECT programme_id FROM programme_phase WHERE id = ?", [id], one=True)
    if not phase_info:
        return jsonify({"error": "Phase introuvable"}), 404

    programme_id = phase_info["programme_id"]
    total = query_db("""
        SELECT SUM(poids) AS total FROM programme_phase
        WHERE programme_id = ? AND id != ?
    """, [programme_id, id], one=True)
    total_poids = (total["total"] or 0) + poids
    if total_poids > 100:
        return jsonify({"error": f"La somme totale ({total_poids:.1f}%) dépasse 100%."}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE programme_phase
        SET poids = ?, udate = DATETIME('now'), uuser = ?
        WHERE id = ?
    """, (poids, session.get('user', {}).get('username', 'admin'), id))
    conn.commit()
    return jsonify({"success": True})


@programme_config_bp.route("/supprimer_phase/<int:id>", methods=["DELETE"])
def supprimer_phase(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM programme_phase WHERE id = ?", (id,))
    conn.commit()
    return jsonify({"success": True})


# ============================================================== #
# 🔹 PROFILS
# ============================================================== #
@programme_config_bp.route("/get_profils/<int:programme_id>")
def get_profils(programme_id):
    profils = query_db("""
        SELECT pp.id, p.nom AS nom_profil, pp.poids
        FROM programme_profils pp
        JOIN Profils p ON p.id = pp.profil_id
        WHERE pp.programme_id = ?
        ORDER BY p.id
    """, [programme_id])
    return jsonify([dict(r) for r in profils])


@programme_config_bp.route("/get_available_profils/<int:programme_id>")
def get_available_profils(programme_id):
    profils = query_db("""
        SELECT p.id, p.nom
        FROM Profils p
        WHERE p.id NOT IN (
            SELECT profil_id FROM programme_profils WHERE programme_id = ?
        )
        ORDER BY p.id
    """, [programme_id])
    return jsonify([dict(r) for r in profils])


@programme_config_bp.route("/ajouter_profil", methods=["POST"])
def ajouter_profil():
    programme_id = request.form.get("programme_id")
    profil_id = request.form.get("profil_id")
    poids = float(request.form.get("poids", 0))

    total = query_db("SELECT SUM(poids) AS total FROM programme_profils WHERE programme_id = ?", [programme_id], one=True)
    total_poids = (total["total"] or 0) + poids
    if total_poids > 100:
        return jsonify({"error": f"La somme totale ({total_poids:.1f}%) dépasse 100%."}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO programme_profils (programme_id, profil_id, poids, iuser)
        VALUES (?, ?, ?, ?)
    """, (programme_id, profil_id, poids, session.get('user', {}).get('username', 'admin')))
    conn.commit()
    return jsonify({"success": True})


@programme_config_bp.route("/modifier_profil/<int:id>", methods=["POST"])
def modifier_profil(id):
    poids = float(request.form.get("poids", 0))
    profil_info = query_db("SELECT programme_id FROM programme_profils WHERE id = ?", [id], one=True)
    if not profil_info:
        return jsonify({"error": "Profil introuvable"}), 404

    programme_id = profil_info["programme_id"]
    total = query_db("""
        SELECT SUM(poids) AS total FROM programme_profils
        WHERE programme_id = ? AND id != ?
    """, [programme_id, id], one=True)
    total_poids = (total["total"] or 0) + poids
    if total_poids > 100:
        return jsonify({"error": f"La somme totale ({total_poids:.1f}%) dépasse 100%."}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE programme_profils
        SET poids = ?, udate = DATETIME('now'), uuser = ?
        WHERE id = ?
    """, (poids, session.get('user', {}).get('username', 'admin'), id))
    conn.commit()
    return jsonify({"success": True})


@programme_config_bp.route("/supprimer_profil/<int:id>", methods=["DELETE"])
def supprimer_profil(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM programme_profils WHERE id = ?", (id,))
    conn.commit()
    return jsonify({"success": True})


# ============================================================== #
# 🔹 PROJETS (UPDATE sur Projet)
# ============================================================== #
@programme_config_bp.route("/get_projets/<int:programme_id>")
def get_projets(programme_id):
    projets = query_db("""
        SELECT id, titre_projet
        FROM Projet
        WHERE id_programme = ?
        ORDER BY id DESC
    """, [programme_id])
    return jsonify([dict(r) for r in projets])


@programme_config_bp.route("/get_available_projets/<int:programme_id>")
def get_available_projets(programme_id):
    projets = query_db("""
        SELECT id, titre_projet
        FROM Projet
        WHERE id_programme IS NULL
           OR id_programme = 0
           OR id_programme = ''
        ORDER BY id DESC
    """)
    return jsonify([dict(r) for r in projets])


@programme_config_bp.route("/ajouter_projet", methods=["POST"])
def ajouter_projet():
    programme_id = request.form.get("programme_id")
    projet_id = request.form.get("projet_id")

    if not programme_id or not projet_id:
        return jsonify({"error": "Paramètres manquants."}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE Projet
        SET id_programme = ?, udate = DATETIME('now'), uuser = ?
        WHERE id = ?
    """, (programme_id, session.get('user', {}).get('username', 'admin'), projet_id))
    conn.commit()
    return jsonify({"success": True})


@programme_config_bp.route("/supprimer_projet/<int:id>", methods=["DELETE"])
def supprimer_projet(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE Projet
        SET id_programme = NULL, udate = DATETIME('now'), uuser = ?
        WHERE id = ?
    """, (session.get('user', {}).get('username', 'admin'), id))
    conn.commit()
    return jsonify({"success": True})


# ==============================================================
# 🔹 TABLEAU DES CHARGES (AJAX)
# ==============================================================
@programme_config_bp.route("/tableau_charges/<int:programme_id>")
def tableau_charges(programme_id):
    try:
        phases = query_db("""
            SELECT f.id, f.nom, ph.poids
            FROM programme_phase ph
            JOIN Phase f ON f.id = ph.phase_id
            WHERE ph.programme_id = ?
            ORDER BY f.id
        """, [programme_id])

        profils = query_db("""
            SELECT p.id, p.nom, pp.poids
            FROM programme_profils pp
            JOIN Profils p ON p.id = pp.profil_id
            WHERE pp.programme_id = ?
            ORDER BY p.id
        """, [programme_id])

        tableau_final = []
        for phase in phases:
            ligne = {"phase": phase["nom"], "poids": phase["poids"], "profils": {}}
            for profil in profils:
                poids_phase = phase["poids"] or 0
                poids_profil = profil["poids"] or 0
                charge = (poids_phase * poids_profil) / 100
                ligne["profils"][profil["id"]] = round(charge, 2)
            tableau_final.append(ligne)

        return render_template(
            "Programme/tableau_charges.html",
            phases=phases,
            profils=profils,
            tableau_final=tableau_final
        )

    except Exception as e:
        print(f"🔥 ERREUR TABLEAU_CHARGES : {e}")
        return render_template(
            "Programme/tableau_charges.html",
            phases=[], profils=[], tableau_final=[]
        )


@programme_config_bp.route("/exporter_tableau_charges/<int:programme_id>")
def exporter_tableau_charges(programme_id):
    # 🔹 Récupération des phases
    phases = query_db("""
        SELECT f.nom AS phase, ph.poids
        FROM programme_phase ph
        JOIN Phase f ON f.id = ph.phase_id
        WHERE ph.programme_id = ?
        ORDER BY f.id
    """, [programme_id])

    # 🔹 Récupération des profils
    profils = query_db("""
        SELECT p.nom, pp.poids, p.id
        FROM programme_profils pp
        JOIN Profils p ON p.id = pp.profil_id
        WHERE pp.programme_id = ?
        ORDER BY p.id
    """, [programme_id])

    if not phases or not profils:
        flash("⚠️ Les phases et/ou profils ne sont pas définis pour ce programme.", "warning")
        return redirect(url_for("programme_config.liste_programmes"))

    # --- Créer un workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau des Charges"

    # --- Styles ---
    blue_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Bleu foncé
    blue_light = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")   # Bleu clair
    gray_footer = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")  # Gris clair
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999")
    )

    # --- En-tête principale ---
    ws.cell(row=1, column=1, value="Profil").fill = blue_header
    ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).border = thin_border

    col = 2
    for phase in phases:
        cell = ws.cell(row=1, column=col, value=phase["phase"])
        cell.fill = blue_header
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        col += 1

    total_col = len(phases) + 2
    ws.cell(row=1, column=total_col - 1, value="Total Profil").fill = blue_header
    ws.cell(row=1, column=total_col - 1).font = Font(color="FFFFFF", bold=True)
    ws.cell(row=1, column=total_col - 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=total_col - 1).border = thin_border

    # --- Contenu principal (profils → lignes) ---
    current_row = 2
    for profil in profils:
        ws.cell(row=current_row, column=1, value=profil["nom"])
        ws.cell(row=current_row, column=1).fill = blue_light
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        total_profil = 0
        for i, phase in enumerate(phases, start=2):
            poids_phase = phase["poids"] or 0
            poids_profil = profil["poids"] or 0
            charge = (poids_phase * poids_profil) / 100
            total_profil += charge

            c = ws.cell(row=current_row, column=i, value=round(charge, 2))
            c.alignment = Alignment(horizontal="center")
            c.fill = white_fill
            c.border = thin_border

        # ✅ Total Profil
        total_cell = ws.cell(row=current_row, column=total_col - 1, value=round(total_profil, 2))
        total_cell.fill = blue_light
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="center")
        total_cell.border = thin_border

        current_row += 1

    # --- Ligne Totaux Phases ---
    ws.cell(row=current_row, column=1, value="Total Phase").fill = gray_footer
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=current_row, column=1).border = thin_border

    for i, phase in enumerate(phases, start=2):
        poids_phase = phase["poids"] or 0
        total_phase = poids_phase  # on affiche directement le poids de la phase
        c = ws.cell(row=current_row, column=i, value=round(total_phase, 2))
        c.fill = gray_footer
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # ✅ Cellule finale = somme 100%
    total_total = ws.cell(row=current_row, column=total_col - 1, value=100)
    total_total.fill = gray_footer
    total_total.font = Font(bold=True, color="1E3A8A")
    total_total.alignment = Alignment(horizontal="center")
    total_total.border = thin_border

    # --- Ajustement automatique des colonnes ---
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    # --- Enregistrer en mémoire et renvoyer ---
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"tableau_charges_programme_{programme_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
